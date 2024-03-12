import openpyxl


class excel_sheet():

    def __init__(self, path: str) -> None:

        self.xlsx_files = path

        self.workbook = openpyxl.load_workbook(self.xlsx_files)
        self.sheet = self.workbook.active
        self.sheet_height = self.sheet.max_row
        self.sheet_width = self.sheet.max_column


    def get_sheet_height(self) -> int:
        return self.sheet_height
    

    def get_sheet_width(self) -> int:
        return self.sheet_width


    #?is there a built-in way to get all the column headers and their indices?
    def get_all_column_headers(self) -> list:
        column_headers = []
        for header in self.sheet.iter_cols(1, self.sheet_width, values_only=True):
            header_text = header[0]
            column_headers.append(header_text)
        return column_headers


    #TODO: figure out a better way of pairing column data-values
    def create_column_data_pair(self, column_0: int, column_1: int) -> list:
        column_data_pair = []
        for row in self.sheet.iter_rows(1, self.sheet_height, values_only=True):
            val0 = row[column_0]
            val1 = row[column_1]
            column_data_pair.append(f"{val0}.{val1}")
        return column_data_pair


    def delete_column(self, columns_to_delete: list) -> None:
        #!NOTE: the 'delete_cols' works as 'in-range' deletion. 
        #!When giving it 2 int args, it will delete the cols in range (example) 10 - 12 will delete 11 aswell
        #!Range 1 - 1 or 'delete_cols(1, 1)' just deletes column A
        for column in columns_to_delete:
            self.sheet.delete_cols(column)


    def save_workbook(self, filename: str) -> None:
        self.workbook.save(filename)


    def close_workbook(self) -> None:
        self.workbook.close()


    def match_and_insert(self, column_lane_index: int, column_insert_index: int, replacement: str, replacement0: str, data: list) -> None:
        row_count = 1
        data_SNs = [i.split(".")[0] for i in data]
        for row in self.sheet.iter_rows(1, self.sheet_height, values_only=True):
            found_cell_value = row[column_lane_index]
            if found_cell_value == None:
                self.sheet.cell(row=row_count, column=column_insert_index, value=replacement)
                row_count += 1
                continue
            if found_cell_value in data_SNs:
                for value in data:
                    data_pair = value.split(".")
                    serial = data_pair[0]
                    po_num = data_pair[1]
                    if serial == found_cell_value:
                        self.sheet.cell(row=row_count, column=column_insert_index, value=po_num)
                        row_count += 1
            else:
                self.sheet.cell(row=row_count, column=column_insert_index, value=replacement0)
                row_count += 1


def main():
    old_sheet_path = "./demo_files/old.xlsx"
    new_sheet_path = "./demo_files/new.xlsx"
    final_sheet_path = "./demo_files/final.xlsx"

    #delete (the unwanted) columns A, K, L ('Supplier parent', 'Goods in', 'Sorted')
    #!when deleting a single column, that moves the other columns down-lane making their index -1 position
    columns_to_delete = [1, 3, 3, 9, 9]

    old_sheet = excel_sheet(old_sheet_path)
    new_sheet = excel_sheet(new_sheet_path)

    new_sheet.delete_column(columns_to_delete)
    new_sheet.save_workbook(final_sheet_path)
    new_sheet.close_workbook()

    final_sheet = excel_sheet(final_sheet_path)

    old_sheet_headers = old_sheet.get_all_column_headers()
    old_sheet_serial_index = old_sheet_headers.index("Serial number")
    old_sheet_po_index = old_sheet_headers.index("PO Kunde")
    old_sheet_sn_po = old_sheet.create_column_data_pair(old_sheet_serial_index, old_sheet_po_index)

    final_sheet_headers = final_sheet.get_all_column_headers()
    final_sheet_serial_index = final_sheet_headers.index("Serial number")

    final_sheet_column_insert_po = 9
    none_to_match_replacement = "HP_ACCESSORIES"
    no_match_replacement = "404_not_found"

    final_sheet.match_and_insert(
        column_lane_index=final_sheet_serial_index, 
        column_insert_index=final_sheet_column_insert_po, 
        replacement=none_to_match_replacement, 
        replacement0=no_match_replacement, 
        data=old_sheet_sn_po
        )

    final_sheet.save_workbook(final_sheet_path)
    final_sheet.close_workbook()
    old_sheet.close_workbook()
    print("\nDONE!\n")


if __name__ == '__main__':
    main()
