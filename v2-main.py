#!./env/bin/python3

import os
import configparser
import openpyxl

class excel_workbook():

    def __init__(self, path: str) -> None:
        self.xlsx_file = path
        self.workbook = openpyxl.load_workbook(self.xlsx_file)
        self.sheet = self.workbook.active
        self.sheet_height = self.sheet.max_row
        self.sheet_width = self.sheet.max_column
        self.blank_cells = 0

    def get_all_column_headers(self) -> list:
        column_headers = []
        for header in self.sheet.iter_cols(1, self.sheet_width, values_only=True):
            column_headers.append(header[0])
        return column_headers

    def map_data_pair(self, column_0: int, column_1: int) -> list:
        column_data_pair = []
        for row in self.sheet.iter_rows(1, self.sheet_height, values_only=True):
            column_data_pair.append(str(row[column_0]) + "." + str(row[column_1]))
        return column_data_pair

    def delete_column(self, columns_delete: list) -> None:
        for column in columns_delete:
            self.sheet.delete_cols(column)

    def save_workbook(self, filename: str) -> None:
        self.workbook.save(filename)

    def close_workbook(self) -> None:
        self.workbook.close() 

    def match_and_insert(self, column_lane_index: int, column_insert_index: int, replacement_0: str, replacement_1: str, data: list) -> int:
        row_count = 1
        data_SNs = [i.split(".")[0] for i in data]
        for row in self.sheet.iter_rows(1, self.sheet_height, values_only=True):
            found_cell_value = row[column_lane_index]
            if found_cell_value == None:
                self.sheet.cell(row=row_count, column=column_insert_index, value=replacement_0)
                row_count += 1
                continue
            if found_cell_value in data_SNs:
                for value in data:
                    data_pair = value.split(".")
                    serial = data_pair[0]
                    po = data_pair[1]
                    if serial == found_cell_value:
                        self.sheet.cell(row=row_count, column=column_insert_index, value=po)
                        row_count += 1
            else:
                self.blank_cells += 1
                self.sheet.cell(row=row_count, column=column_insert_index, value=replacement_1)
                row_count += 1

    def get_blank_cells(self) -> int:
        return self.blank_cells

def date_is_valid(date: str) -> bool:
    days = int(date[0:2])
    months = int(date[3:5])
    years = int(date[6:])
    if (days <= 31) and (months <= 12) and ((years < 2100) and (years > 1970)) :
        return True
    else:
        return False

def get_date_from_str(file_name: str) -> str | None:
    char_list = ".0123456789" #instead of trying to convert every character to int
    date_format = "dd.mm.yyyy"
    for index, char in enumerate(file_name):
        if char in char_list:
            date_format_lenght = len(date_format) + 1
            for x in range(date_format_lenght):
                x_val = file_name[index + x]
                if (x == 2 or x == 5) and (x_val != '.'):
                    break
                elif x == (date_format_lenght - 1):
                    date_found = file_name[index:(index + x)]
                    if date_is_valid(date_found):
                        return date_found
                    else:
                        break
                elif x_val not in char_list:
                    break
                else:
                    continue
        else:
            continue

# THIS ONE IS BIG BRAIN UNIQUE BRO.
def get_newest_date(d1: str, d2: str) -> str | None:
    if d1 == d2:
        return None
    d1_formated = d1.split('.')[::-1]
    d2_formated = d2.split('.')[::-1]
    if d1_formated > d2_formated:
        return d1
    else:
        return d2

def get_excel_files() -> list:
    files = os.listdir('.')
    excel_files = []
    for file in files:
        if file[-5:] == '.xlsx':
            excel_files.append(file)
    return excel_files

def main() -> None:
    # set configparser to read the config.
    config = configparser.ConfigParser()
    config.read('config.ini')
    default_config = config['DEFAULT']

    # set all config.ini-values to their respective variables, making them usable.
    old_sheet_path = default_config['old_sheet_path']
    new_sheet_path = default_config['new_sheet_path']
    final_sheet_path = default_config['final_sheet_path']
    final_sheet_name = default_config['final_sheet_name']
    columns_to_delete = default_config['columns_to_delete']
    serial_column_text = default_config['serial_column_text']
    po_column_text = default_config['po_column_text']
    final_sheet_column_insert_po = default_config['final_sheet_column_insert_po']
    none_to_match_replacement = default_config['none_to_match_replacement']
    no_match_replacement = default_config['no_match_replacement']

    # converted to int, since configparser only keeps strings.
    final_sheet_column_insert_po = int(final_sheet_column_insert_po)

    # find the files to be manipulated.
    excel_files = get_excel_files()
  
    # assign the excel files to variables based on the dates detected in their names, NOT meta-data dates.
    new_sheet_name = str(get_newest_date(excel_files[0], excel_files[1]))
    if new_sheet_name == excel_files[0]:
        old_sheet_name = excel_files[1]
    else:
        old_sheet_name = excel_files[0]

    #TODO: implement a warning-function for when things dont go as planned (eg: multiple xlsx files in the the dir, the dates of files are the same, etc)

    # assign the full file-path based on previous variables, to other variables that combine the path and the file-name.
    new_sheet_path = new_sheet_path + new_sheet_name
    old_sheet_path = old_sheet_path + old_sheet_name
    final_sheet_path = final_sheet_path + final_sheet_name

    # activate the workbooks.
    old_sheet = excel_workbook(old_sheet_path)
    new_sheet = excel_workbook(new_sheet_path)
    
    # need to convert "columns_to_delete" from string into list, for the compability.
    columns_to_delete = columns_to_delete.split(',')
    # use eval to convert every str-number to actual int.
    columns_to_delete = [eval(i) for i in columns_to_delete]
    new_sheet.delete_column(columns_to_delete)
    new_sheet.save_workbook(final_sheet_path)
    new_sheet.close_workbook()

    final_sheet = excel_workbook(final_sheet_path)

    old_sheet_headers = old_sheet.get_all_column_headers()
    old_sheet_serial_index = old_sheet_headers.index(serial_column_text)
    old_sheet_po_index = old_sheet_headers.index(po_column_text)
    old_sheet_sn_po = old_sheet.map_data_pair(old_sheet_serial_index, old_sheet_po_index)

    final_sheet_headers = final_sheet.get_all_column_headers()
    final_sheet_serial_index = final_sheet_headers.index(serial_column_text)
    
    final_sheet.match_and_insert(
            column_lane_index=final_sheet_serial_index,
            column_insert_index=final_sheet_column_insert_po,
            replacement_0=none_to_match_replacement,
            replacement_1=no_match_replacement,
            data=old_sheet_sn_po
            )

    final_sheet_blank_POs = final_sheet.get_blank_cells()
    if final_sheet_blank_POs > 0:
        print(f"[!] Serial numbers without PO: {final_sheet_blank_POs}")
    
    final_sheet.save_workbook(final_sheet_path)
    final_sheet.close_workbook()
    old_sheet.close_workbook()
    print("\n[+] DONE!\n")

    input("[x] Press [ENTER] to exit the application")

if __name__ == '__main__':
    main()
